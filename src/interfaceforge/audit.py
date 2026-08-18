"""Mode-aware, dependency-light audit of VASP and VASP-MLFF run folders."""

from __future__ import annotations

import csv
import json
import math
import re
import time
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

from .state import StateStore
from .vasp import parse_incar

RUN_MARKERS = ("INCAR", "OUTCAR", "OSZICAR", "ML_LOGFILE", "ML_AB", "ML_ABN", "ML_FF", "ML_FFN")
MAX_TEXT_BYTES = 80 * 1024 * 1024

# Heuristic health thresholds used by assess(). These are conservative
# defaults for typical VASP-MLFF on-the-fly campaigns, not values derived
# from a specific validation study; treat them as a starting point that may
# need site- or system-specific tuning rather than a scientific ground truth.
LEARNING_RATE_HEAVY_PCT = 40.0
LEARNING_RATE_MODERATE_PCT = 15.0
LEARNING_RATE_NEAR_REFIT_PCT = 3.0
READINESS_PROFILES = ("general", "perovskite")
PEROVSKITE_RECENT_WINDOW = 250
PEROVSKITE_MIN_PLATEAU_RECORDS = 200
PEROVSKITE_MAX_RECENT_LEARNING_PCT = 20.0
PEROVSKITE_MAX_BEEF_P95_EV_A = 0.03
PEROVSKITE_MAX_BEEF_HALF_DELTA_EV_A = 0.002
SFF_STRONG_EXTRAPOLATION = 0.8
SFF_SUBSTANTIAL_EXTRAPOLATION = 0.2
SFF_MODERATE_EXTRAPOLATION = 0.05

AT_A_GLANCE_COLUMNS = (
    ("run", "Run"),
    ("relative_path", "Relative path"),
    ("readiness_profile", "Readiness profile"),
    ("ml_mode", "Mode"),
    ("run_kind", "Run kind"),
    ("progress_pct", "Progress (%)"),
    ("md_steps", "MD steps"),
    ("ionic_steps", "Ionic steps"),
    ("nsw_target", "NSW target"),
    ("opt_converged", "OPT converged"),
    ("sigma0_energy_ev_last", "Energy sigma->0 last (eV)"),
    ("sigma0_energy_delta_last_ev", "Energy delta last (eV)"),
    ("max_force_ev_a_last", "Max force last (eV/A)"),
    ("temperature_mean_k", "MD mean T (K)"),
    ("temperature_std_k", "MD std T (K)"),
    ("finished_normally", "Finished normally"),
    ("health", "Health"),
    ("learning_event_rate_pct", "Learning events (%)"),
    ("recent_learning_event_rate_pct", "Recent learning events (%)"),
    ("critical_events", "Critical events"),
    ("recent_critical_events", "Recent critical events"),
    ("training_force_rmse_ev_a_last", "Training force RMSE (eV/A)"),
    ("bayesian_force_error_ev_a_last", "BEEF last (eV/A)"),
    ("bayesian_force_error_ev_a_max", "BEEF max (eV/A)"),
    ("ml_ctifor_ev_a_last", "ML_CTIFOR last (eV/A)"),
    ("recent_beef_p95_ev_a", "Recent BEEF p95 (eV/A)"),
    ("recent_beef_half_delta_ev_a", "Recent BEEF half-delta (eV/A)"),
    ("perovskite_sampling_plateau", "Perovskite plateau"),
    ("true_force_error_ev_a_last", "True force error last (eV/A)"),
    ("sff_max_atom_max", "SFF max"),
    ("lconf_max", "LCONF max"),
    ("warnings", "Warnings"),
    ("next_action", "Next action"),
)


def read_tail(path: Path, max_bytes: int = MAX_TEXT_BYTES) -> str:
    if not path.is_file():
        return ""
    try:
        size = path.stat().st_size
        with path.open("rb") as handle:
            if size > max_bytes:
                handle.seek(size - max_bytes)
            return handle.read().decode("utf-8", errors="ignore")
    except OSError:
        return ""


def _number(tokens: list[str], index: int) -> float | None:
    if index >= len(tokens):
        return None
    try:
        return float(tokens[index])
    except ValueError:
        return None


def _stats(values: list[float], prefix: str) -> dict[str, float | None]:
    return {
        f"{prefix}_last": values[-1] if values else None,
        f"{prefix}_mean": mean(values) if values else None,
        f"{prefix}_max": max(values) if values else None,
    }


def _percentile(values: list[float], fraction: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    position = (len(ordered) - 1) * fraction
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1.0 - weight) + ordered[upper] * weight


def parse_ml_log(path: Path) -> dict[str, Any]:
    status = learning = critical = 0
    status_events: list[tuple[bool, bool]] = []
    err_energy: list[float] = []
    err_force: list[float] = []
    err_stress: list[float] = []
    beef: list[float] = []
    threshold: list[float] = []
    spilling: list[float] = []
    lconf: list[int] = []
    text = read_tail(path)
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        tokens = stripped.split()
        record = tokens[0].upper()
        upper = stripped.upper()
        if "STATUS" in upper:
            status += 1
            is_learning = bool(re.search(r"\bLEARN(?:ING)?\b", upper))
            is_critical = bool(re.search(r"\bCRIT(?:ICAL)?\b", upper))
            learning += is_learning
            critical += is_critical
            status_events.append((is_learning, is_critical))
        if record == "ERR":
            if (value := _number(tokens, 2)) is not None:
                err_energy.append(value)
            if (value := _number(tokens, 3)) is not None:
                err_force.append(value)
            if (value := _number(tokens, 4)) is not None:
                err_stress.append(value)
        elif record in {"BEEF", "BEFF"}:
            if (value := _number(tokens, 3)) is not None:
                beef.append(value)
            if (value := _number(tokens, 5)) is not None:
                threshold.append(value)
        elif record == "SFF" and (value := _number(tokens, 2)) is not None:
            spilling.append(value)
        elif record == "LCONF":
            integers = [int(token) for token in tokens[1:] if re.fullmatch(r"[-+]?\d+", token)]
            if integers:
                lconf.append(integers[-1])
    recent_status = status_events[-PEROVSKITE_RECENT_WINDOW:]
    recent_beef = beef[-PEROVSKITE_RECENT_WINDOW:]
    midpoint = len(recent_beef) // 2
    recent_beef_half_delta = (
        mean(recent_beef[midpoint:]) - mean(recent_beef[:midpoint])
        if midpoint and len(recent_beef[midpoint:])
        else None
    )
    recent_learning = sum(event[0] for event in recent_status)
    recent_critical = sum(event[1] for event in recent_status)
    recent_learning_rate = (
        100.0 * recent_learning / len(recent_status) if recent_status else None
    )
    recent_beef_p95 = _percentile(recent_beef, 0.95)
    perovskite_plateau = (
        len(recent_beef) >= PEROVSKITE_MIN_PLATEAU_RECORDS
        and recent_beef_p95 is not None
        and recent_beef_p95 <= PEROVSKITE_MAX_BEEF_P95_EV_A
        and recent_beef_half_delta is not None
        and abs(recent_beef_half_delta) <= PEROVSKITE_MAX_BEEF_HALF_DELTA_EV_A
        and recent_critical == 0
        and recent_learning_rate is not None
        and recent_learning_rate <= PEROVSKITE_MAX_RECENT_LEARNING_PCT
    )
    result: dict[str, Any] = {
        "ml_log_present": bool(text),
        "status_records": status,
        "learning_events": learning,
        "critical_events": critical,
        "learning_event_rate_pct": 100.0 * learning / status if status else None,
        "recent_status_records": len(recent_status),
        "recent_learning_events": recent_learning,
        "recent_learning_event_rate_pct": recent_learning_rate,
        "recent_critical_events": recent_critical,
        "recent_beef_records": len(recent_beef),
        "recent_beef_p95_ev_a": recent_beef_p95,
        "recent_beef_half_delta_ev_a": recent_beef_half_delta,
        "perovskite_sampling_plateau": perovskite_plateau,
        "training_rmse_records": len(err_force),
        "lconf_last": lconf[-1] if lconf else None,
        "lconf_max": max(lconf) if lconf else None,
    }
    result.update(_stats(err_energy, "training_energy_rmse_ev_atom"))
    result.update(_stats(err_force, "training_force_rmse_ev_a"))
    result.update(_stats(err_stress, "training_stress_rmse_kbar"))
    # Compatibility alias retained for audit consumers written before the ERR
    # quantity was labeled explicitly as a training-set RMSE.
    result.update(_stats(err_force, "true_force_error_ev_a"))
    result.update(_stats(beef, "bayesian_force_error_ev_a"))
    result.update(_stats(threshold, "ml_ctifor_ev_a"))
    result.update(_stats(spilling, "sff_max_atom"))
    return result


def parse_oszicar(path: Path) -> dict[str, Any]:
    temperatures: list[float] = []
    last_step: int | None = None
    for line in read_tail(path).splitlines():
        match = re.search(r"\bT=\s*([-+0-9.Ee]+)", line)
        if not match:
            continue
        try:
            temperatures.append(float(match.group(1)))
        except ValueError:
            continue
        step_match = re.match(r"^\s*(\d+)\s+", line)
        if step_match:
            last_step = int(step_match.group(1))
    return {
        "md_steps": len(temperatures),
        "last_oszicar_step": last_step,
        "temperature_last_k": temperatures[-1] if temperatures else None,
        "temperature_mean_k": mean(temperatures) if temperatures else None,
        "temperature_std_k": pstdev(temperatures) if len(temperatures) > 1 else None,
        "temperature_min_k": min(temperatures) if temperatures else None,
        "temperature_max_k": max(temperatures) if temperatures else None,
    }


_SIGMA0_ENERGY = re.compile(r"energy\(sigma->0\)\s*=\s*([-+0-9.Ee]+)")
_WITHOUT_ENTROPY_ENERGY = re.compile(r"energy\s+without\s+entropy\s*=\s*([-+0-9.Ee]+)")
_OPT_CONVERGED_MARKER = "reached required accuracy - stopping structural energy minimisation"
_TOTAL_FORCE_BLOCK = re.compile(r"TOTAL-FORCE \(eV/Angst\)\s*\n\s*-+\n(.*?)\n\s*-+", re.S)


def parse_relaxation_progress(run: Path) -> dict[str, Any]:
    """Track ionic-step energy and force progress for a standard (non-ML) run.

    ``energy(sigma->0)`` is the same quantity historically collected across
    many run folders with a one-line ``grep ... | tail -1 | awk '{print $7}'``
    pipeline against "FREE ENERGIE OF THE ION-ELECTRON SYSTEM"; this keeps
    every ionic step so a convergence trend, not just the final value, is
    available.
    """

    # Normalize CRLF up front: the other parsers in this module work
    # line-by-line via str.splitlines(), which treats "\r\n" as one line
    # break automatically, but the TOTAL-FORCE block below matches literal
    # "\n" across multiple lines and would otherwise silently miss every
    # block in a CRLF-terminated OUTCAR.
    text = read_tail(run / "OUTCAR").replace("\r\n", "\n")
    sigma0 = [float(value) for value in _SIGMA0_ENERGY.findall(text)]
    without_entropy = [float(value) for value in _WITHOUT_ENTROPY_ENERGY.findall(text)]
    forces: list[float] = []
    for block in _TOTAL_FORCE_BLOCK.findall(text):
        magnitudes = []
        for line in block.strip().splitlines():
            tokens = line.split()
            if len(tokens) < 6:
                continue
            try:
                fx, fy, fz = (float(tokens[-3]), float(tokens[-2]), float(tokens[-1]))
            except ValueError:
                continue
            magnitudes.append(math.sqrt(fx * fx + fy * fy + fz * fz))
        if magnitudes:
            forces.append(max(magnitudes))
    result: dict[str, Any] = {
        "ionic_steps": len(sigma0),
        "opt_converged": _OPT_CONVERGED_MARKER in text.lower(),
        "max_force_ev_a_last": forces[-1] if forces else None,
        "max_force_ev_a_max": max(forces) if forces else None,
        "sigma0_energy_delta_last_ev": sigma0[-1] - sigma0[-2] if len(sigma0) >= 2 else None,
    }
    result.update(_stats(sigma0, "sigma0_energy_ev"))
    result.update(_stats(without_entropy, "energy_without_entropy_ev"))
    return result


def classify_run_kind(row: dict[str, Any], incar: dict[str, str]) -> str:
    """Classify a run as an MLFF stage, a standard relaxation, MD, or static.

    ML runs are already distinguished by ``ML_MODE``/``ML_LMLFF``. For
    everything else, InterfaceForge previously had no opinion beyond
    "unknown mode"; this reads the same ``IBRION``/``NSW`` tags a human would
    to tell a geometry optimization apart from ab initio MD or a single
    static point.
    """

    if row.get("ml_mode") or str(row.get("ml_lmlff", "")).strip().upper() in {"T", ".TRUE.", "TRUE"}:
        return f"ml_{row.get('ml_mode') or 'unknown'}"
    ibrion_raw = incar.get("IBRION")
    try:
        ibrion = int(float(ibrion_raw)) if ibrion_raw not in (None, "") else None
    except ValueError:
        ibrion = None
    nsw = row.get("nsw_target") or 0
    if ibrion in (1, 2, 3, 5, 6, 7, 8) and nsw > 0:
        return "opt"
    if ibrion == 0 and nsw > 0:
        return "md"
    return "static"


def _assess_opt(row: dict[str, Any]) -> Assessment:
    if row.get("oom_or_killed"):
        return Assessment("failed: OOM/killed", "Fix memory or resources before interpreting the run.")
    if not row.get("ionic_steps"):
        return Assessment(
            "no ionic steps parsed",
            "Inspect OUTCAR for an initialization or SCF-convergence failure.",
        )
    if row.get("opt_converged"):
        return Assessment(
            "converged",
            "Relaxation reached VASP's required accuracy. Inspect CONTCAR, then "
            "consider a follow-up static run for final energies/properties.",
        )
    if not row.get("finished_normally"):
        return Assessment("incomplete or running", "Inspect scheduler state and the latest ionic step in OUTCAR.")
    target = row.get("nsw_target")
    steps = row.get("ionic_steps") or 0
    if target and steps >= target:
        return Assessment(
            "not converged: reached NSW",
            "Continue from CONTCAR with additional NSW, or reconsider EDIFFG/ISIF "
            "if the geometry is not expected to move further.",
        )
    delta = row.get("sigma0_energy_delta_last_ev")
    if delta is not None and delta > 0:
        return Assessment(
            "energy increased on the last ionic step",
            "Inspect the last few ionic steps and forces; consider restarting "
            "from an earlier CONTCAR with a smaller POTIM or a different IBRION.",
        )
    return Assessment(
        "finished; convergence marker not found",
        "Inspect the final OUTCAR ionic step and forces before accepting CONTCAR.",
    )


def _assess_md(row: dict[str, Any]) -> Assessment:
    if row.get("oom_or_killed"):
        return Assessment("failed: OOM/killed", "Fix memory or resources before interpreting the run.")
    if not row.get("md_steps"):
        return Assessment("no MD steps parsed", "Fix initialization or SCF convergence.")
    if not row.get("finished_normally"):
        return Assessment("incomplete or running", "Inspect scheduler state and the latest OSZICAR step.")
    tebeg = row.get("tebeg_k")
    teend = row.get("teend_k")
    target = (float(tebeg) + float(teend)) / 2 if tebeg is not None and teend is not None else None
    mean_temp = row.get("temperature_mean_k")
    if target and mean_temp is not None and abs(mean_temp - target) > 0.15 * target:
        return Assessment(
            "temperature drift from target",
            f"Mean temperature {mean_temp:.1f} K deviates from the ~{target:.0f} K "
            "target by more than 15%; check the thermostat (MDALGO/SMASS) and "
            "equilibration length.",
        )
    return Assessment(
        "completed; average behavior reported",
        "Review temperature_mean_k/temperature_std_k and inspect XDATCAR for "
        "structural drift before using these frames.",
    )


def parse_outcar(run: Path) -> dict[str, Any]:
    chunks = [read_tail(run / "OUTCAR")]
    for pattern in ("slurm-*.out", "slurm-*.err", "vasp.*.out", "vasp.*.err"):
        chunks.extend(read_tail(path, 10 * 1024 * 1024) for path in sorted(run.glob(pattern))[-3:])
    text = "\n".join(chunks)
    warnings: list[str] = []
    patterns = (
        ("OOM/killed", r"out of memory|\boom\b|cannot allocate memory|killed"),
        (
            "ML local-reference capacity",
            r"not enough storage reserved for local reference configurations|"
            r"increase\s+ML_M(?:B|CONF)|ML_M(?:B|CONF).*(?:too small|exceed)",
        ),
        ("BRMIX", r"\bBRMIX\b"),
        ("ZBRENT", r"\bZBRENT\b"),
        ("serious problems", r"very serious problems"),
        ("ML convergence", r"convergence problem in md"),
    )
    for label, pattern in patterns:
        if re.search(pattern, text, re.I):
            warnings.append(label)
    return {
        "finished_normally": "general timing and accounting informations" in text.lower(),
        "oom_or_killed": "OOM/killed" in warnings,
        "ml_capacity_stop": "ML local-reference capacity" in warnings,
        "warnings": "; ".join(warnings),
    }


def _file_mb(path: Path) -> float | None:
    return path.stat().st_size / (1024 * 1024) if path.is_file() else None


def _fast_header(run: Path) -> str:
    for name in ("ML_FFN", "ML_FF"):
        path = run / name
        if not path.is_file():
            continue
        header = path.open("rb").read(4096).decode("ascii", errors="ignore")
        match = re.search(r'"?ML_LFAST"?\s*[:=]\s*(true|false|T|F)', header, re.I)
        if match:
            return match.group(1).lower()
    return ""


@dataclass(frozen=True)
class Assessment:
    health: str
    next_action: str


def _assess_perovskite_training(row: dict[str, Any]) -> Assessment:
    """Assess a fluxional perovskite sampling stage without demanding zero learning events."""

    plateau = bool(row.get("perovskite_sampling_plateau"))
    capacity_stop = bool(row.get("ml_capacity_stop"))
    if plateau:
        if capacity_stop:
            health = "perovskite sampling checkpoint reached"
        elif row.get("finished_normally"):
            health = "perovskite sampling budget complete"
        else:
            health = "perovskite sampling plateau reached"
        return Assessment(
            health,
            "Preserve ML_ABN; stop at a safe checkpoint, then SELECT/REFIT and "
            "validate with prediction MD and held-out DFT.",
        )
    if capacity_stop:
        return Assessment(
            "perovskite sampling incomplete: capacity stop",
            "Recent sampling has not met the perovskite plateau checks; continue "
            "targeted sampling or expand/reselect deliberately.",
        )
    if not row.get("ml_log_present"):
        return Assessment("failed or not initialized", "Inspect MLFF initialization and scheduler output.")
    if not row.get("md_steps"):
        return Assessment("no MD steps parsed", "Fix initialization or SCF convergence.")
    if row.get("finished_normally"):
        return Assessment(
            "perovskite sampling budget complete; validation required",
            "Do not extend solely to reduce learning events; SELECT/REFIT, run "
            "prediction MD, and validate against held-out DFT.",
        )
    if int(row.get("recent_beef_records") or 0) < PEROVSKITE_MIN_PLATEAU_RECORDS:
        return Assessment(
            "perovskite sampling window incomplete",
            f"Collect at least {PEROVSKITE_MIN_PLATEAU_RECORDS} usable BEEF records "
            "before applying the plateau checkpoint.",
        )
    if int(row.get("recent_critical_events") or 0) > 0:
        return Assessment(
            "perovskite sampling still critical",
            "Continue focused training until the recent window contains no critical events.",
        )
    if float(row.get("recent_learning_event_rate_pct") or 0.0) > PEROVSKITE_MAX_RECENT_LEARNING_PCT:
        return Assessment(
            "perovskite sampling still broadening",
            "Continue focused sampling; the recent learning-event rate remains above the perovskite profile limit.",
        )
    return Assessment(
        "perovskite BEEF not stationary",
        "Continue only until the recent BEEF tail and half-window drift stabilize, then SELECT/REFIT and validate.",
    )


def assess(row: dict[str, Any], readiness_profile: str = "general") -> Assessment:
    mode = str(row.get("ml_mode", "")).lower()
    if row["oom_or_killed"]:
        return Assessment("failed: OOM/killed", "Fix memory or resources before interpreting the run.")
    if mode == "train" and readiness_profile == "perovskite":
        return _assess_perovskite_training(row)
    if row.get("ml_capacity_stop"):
        return Assessment(
            "stopped: ML local-reference capacity",
            "Resume with bounded-memory local-basis discarding, or explicitly increase ML_MB.",
        )
    if mode == "train":
        if not row["ml_log_present"]:
            return Assessment("failed or not initialized", "Inspect MLFF initialization and scheduler output.")
        if not row["md_steps"]:
            return Assessment("no MD steps parsed", "Fix initialization or SCF convergence.")
        if not row["finished_normally"]:
            return Assessment("incomplete or running", "Inspect scheduler state and the latest output.")
        rate = row["learning_event_rate_pct"]
        if rate is None:
            return Assessment("training complete; status unclear", "Inspect STATUS records, then refit.")
        if rate >= LEARNING_RATE_HEAVY_PCT:
            return Assessment("still learning heavily", "Continue focused on-the-fly training.")
        if rate >= LEARNING_RATE_MODERATE_PCT:
            return Assessment("moderately undertrained", "Run one or two focused continuation stages.")
        if rate >= LEARNING_RATE_NEAR_REFIT_PCT:
            return Assessment("near refit/validation", "Consider one continuation, then refit and validate.")
        return Assessment("ready to refit and test", "Refit, run stability MD, and validate against held-out DFT.")
    if mode == "refit":
        if not row["finished_normally"]:
            return Assessment("refit incomplete or failed", "Inspect OUTCAR and rerun refit.")
        if not row.get("ml_ffn_size_mb"):
            return Assessment("refit output missing", "No nonempty ML_FFN was produced.")
        if row.get("ml_lfast_header") not in {"true", "t"}:
            return Assessment("refit completed but not fast", "Check ML_FFN and refit settings.")
        return Assessment("fast refit completed", "Use ML_FFN as ML_FF for stability and validation.")
    if mode == "run":
        if not row["md_steps"] and int(row.get("nsw_target") or 0) != 0:
            return Assessment("prediction MD not started", "Check ML_FF compatibility and initialization.")
        if not row["finished_normally"]:
            return Assessment("prediction incomplete or running", "Inspect scheduler state and geometry.")
        spilling = row.get("sff_max_atom_max")
        if spilling is None:
            return Assessment("run completed; no SFF", "Enable ML_ESTBLOCK and perform held-out validation.")
        if spilling >= SFF_STRONG_EXTRAPOLATION:
            return Assessment("strong extrapolation", "Stop before breakdown and relabel representative frames.")
        if spilling >= SFF_SUBSTANTIAL_EXTRAPOLATION:
            return Assessment("substantial extrapolation", "Inspect and relabel high-SFF frames.")
        if spilling >= SFF_MODERATE_EXTRAPOLATION:
            return Assessment("moderate extrapolation", "Prioritize high-SFF frames for DFT checks.")
        return Assessment("stable within represented space", "Proceed to paired DFT/ML validation.")
    if not mode:
        run_kind = row.get("run_kind")
        if run_kind == "opt":
            return _assess_opt(row)
        if run_kind == "md":
            return _assess_md(row)
        if run_kind == "static":
            if row.get("finished_normally"):
                return Assessment("static calculation finished", "Inspect OUTCAR for the desired properties.")
            return Assessment("incomplete or running", "Inspect scheduler state and the latest OUTCAR output.")
    return Assessment("unknown mode", "Set or inspect ML_MODE; supported modes are train, refit, and run.")


def _excluded(path: Path, root: Path, *, include_archives: bool) -> bool:
    """Return whether a descendant should be omitted from run discovery.

    Archive filtering is evaluated relative to the requested audit root. This
    prevents an unrelated ancestor such as ``/project/archive_storage`` from
    hiding an otherwise valid campaign. The root itself is always eligible
    because passing an archive directory directly is an explicit request.
    """

    if path == root:
        return False
    try:
        parts = path.relative_to(root).parts
    except ValueError:
        parts = path.parts
    lowered = tuple(part.lower() for part in parts)
    if ".interfaceforge" in lowered:
        interfaceforge_index = lowered.index(".interfaceforge")
        recovery_archive = "archive" in lowered[interfaceforge_index + 1 :]
        if not (include_archives and recovery_archive):
            return True
    return not include_archives and any("archive" in part for part in lowered)


def find_runs(
    root: str | Path,
    recursive: bool = True,
    *,
    include_archives: bool = False,
) -> list[Path]:
    campaign_root = Path(root).resolve()
    candidates: Iterable[Path] = campaign_root.rglob("*") if recursive else campaign_root.iterdir()
    runs: set[Path] = set()
    if any((campaign_root / marker).exists() for marker in RUN_MARKERS):
        runs.add(campaign_root)
    for path in candidates:
        if (
            path.is_dir()
            and not _excluded(path, campaign_root, include_archives=include_archives)
            and any((path / marker).exists() for marker in RUN_MARKERS)
        ):
            runs.add(path)
    return sorted(runs)


def audit_run(run: Path, root: Path, *, readiness_profile: str = "general") -> dict[str, Any]:
    if readiness_profile not in READINESS_PROFILES:
        raise ValueError(
            f"Unknown readiness profile {readiness_profile!r}; choose one of {', '.join(READINESS_PROFILES)}"
        )
    incar = parse_incar(run / "INCAR")
    row: dict[str, Any] = {
        "run": run.name,
        "relative_path": "." if run == root else str(run.relative_to(root)),
        "path": str(run),
        "readiness_profile": readiness_profile,
        "ml_mode": incar.get("ML_MODE", ""),
        "ml_lmlff": incar.get("ML_LMLFF", ""),
        "ml_lbasis_discard": incar.get("ML_LBASIS_DISCARD", ""),
        "ml_eps_low": incar.get("ML_EPS_LOW", ""),
        "ml_ialgo_linreg": incar.get("ML_IALGO_LINREG", ""),
        "ml_sion1": incar.get("ML_SION1", ""),
        "ml_mrb2": incar.get("ML_MRB2", ""),
        "nsw_target": int(float(incar["NSW"])) if incar.get("NSW") else None,
        "potim_fs": float(incar["POTIM"]) if incar.get("POTIM") else None,
        "tebeg_k": float(incar["TEBEG"]) if incar.get("TEBEG") else None,
        "teend_k": float(incar["TEEND"]) if incar.get("TEEND") else None,
        "ml_estblock": int(float(incar.get("ML_ESTBLOCK", incar.get("ML_IERR", "0")))),
        "deprecated_ml_istart": "ML_ISTART" in incar,
    }
    row.update(parse_outcar(run))
    row.update(parse_oszicar(run / "OSZICAR"))
    row.update(parse_ml_log(run / "ML_LOGFILE"))
    row.update(parse_relaxation_progress(run))
    row["run_kind"] = classify_run_kind(row, incar)
    for name in ("ML_AB", "ML_ABN", "ML_FF", "ML_FFN", "CONTCAR", "XDATCAR"):
        row[f"{name.lower()}_size_mb"] = _file_mb(run / name)
    row["ml_lfast_header"] = _fast_header(run)
    updates = [
        path
        for name in ("OUTCAR", "OSZICAR", "ML_LOGFILE", "XDATCAR", "CONTCAR")
        if (path := run / name).is_file()
    ]
    if updates:
        newest = max(updates, key=lambda item: item.stat().st_mtime)
        row["last_updated_file"] = newest.name
        row["last_update_age_min"] = max(0.0, time.time() - newest.stat().st_mtime) / 60
    else:
        row["last_updated_file"] = ""
        row["last_update_age_min"] = None
    target = row.get("nsw_target")
    steps_for_progress = row["ionic_steps"] if row["run_kind"] == "opt" else row["md_steps"]
    row["progress_pct"] = 100.0 * steps_for_progress / target if target else None
    verdict = assess(row, readiness_profile)
    row["health"] = verdict.health
    row["next_action"] = verdict.next_action
    return row


def _clean(value: Any) -> Any:
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _at_a_glance_row(row: dict[str, Any]) -> dict[str, Any]:
    return {label: row.get(key) for key, label in AT_A_GLANCE_COLUMNS}


def _write_workbook(
    path: Path,
    summary_rows: list[dict[str, Any]],
    full_rows: list[dict[str, Any]],
) -> bool:
    """Write a two-sheet Excel audit when the optional report dependency exists."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "At a glance"
    full_sheet = workbook.create_sheet("Full audit")

    def populate(sheet: Any, rows: list[dict[str, Any]], fields: list[str], table_name: str) -> None:
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field) for field in fields])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column]
            width = min(max(max(map(len, values), default=0) + 2, 10), 48)
            sheet.column_dimensions[column[0].column_letter].width = width
        if rows:
            table = Table(displayName=table_name, ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

    summary_fields = [label for _, label in AT_A_GLANCE_COLUMNS]
    full_fields = list(dict.fromkeys(key for row in full_rows for key in row)) or ["run"]
    populate(summary_sheet, summary_rows, summary_fields, "AuditAtAGlance")
    populate(full_sheet, full_rows, full_fields, "AuditFull")
    workbook.save(path)
    return True


def run_audit(
    root: str | Path,
    *,
    output_dir: str | Path | None = None,
    recursive: bool = True,
    include_archives: bool = False,
    readiness_profile: str = "general",
) -> dict[str, Any]:
    """Audit all detected runs and write JSON, CSV, and Markdown summaries."""

    if readiness_profile not in READINESS_PROFILES:
        raise ValueError(
            f"Unknown readiness profile {readiness_profile!r}; choose one of "
            f"{', '.join(READINESS_PROFILES)}"
        )
    campaign_root = Path(root).resolve()
    destination = Path(output_dir).resolve() if output_dir else campaign_root / "reports" / "audit"
    destination.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            key: _clean(value)
            for key, value in audit_run(
                run,
                campaign_root,
                readiness_profile=readiness_profile,
            ).items()
        }
        for run in find_runs(
            campaign_root,
            recursive,
            include_archives=include_archives,
        )
    ]
    counts = Counter(row["health"] for row in rows)
    payload = {
        "schema_version": 1,
        "root": str(campaign_root),
        "readiness_profile": readiness_profile,
        "run_count": len(rows),
        "health_counts": dict(counts),
        "runs": rows,
    }

    json_path = destination / "audit.json"
    csv_path = destination / "audit.csv"
    summary_csv_path = destination / "audit_summary.csv"
    workbook_path = destination / "audit.xlsx"
    md_path = destination / "audit.md"
    json_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    fields = list(dict.fromkeys(key for row in rows for key in row))
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields or ["run"])
        writer.writeheader()
        writer.writerows(rows)
    summary_rows = [_at_a_glance_row(row) for row in rows]
    summary_fields = [label for _, label in AT_A_GLANCE_COLUMNS]
    with summary_csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=summary_fields)
        writer.writeheader()
        writer.writerows(summary_rows)
    workbook_written = _write_workbook(workbook_path, summary_rows, rows)
    markdown = [
        "# InterfaceForge VASP audit\n",
        f"Readiness profile: **{readiness_profile}**.\n",
        f"Audited **{len(rows)}** run folder(s).\n",
    ]
    if counts:
        markdown.extend(
            [
                "## Health counts\n",
                *[f"- **{name}**: {count}" for name, count in sorted(counts.items())],
                "",
            ]
        )
    if rows:
        markdown.extend(
            [
                "## Runs\n",
                "| Run | Mode | Progress | Force RMSE (train) | Health | Next action |",
                "|---|---:|---:|---:|---|---|",
            ]
        )
        for row in rows:
            progress = "" if row["progress_pct"] is None else f"{row['progress_pct']:.1f}%"
            force_rmse = row.get("training_force_rmse_ev_a_last")
            force_rmse_text = "" if force_rmse is None else f"{force_rmse:.6g} eV/A"
            mode_text = row["ml_mode"] or row.get("run_kind", "")
            markdown.append(
                f"| `{row['relative_path']}` | {mode_text} | {progress} | {force_rmse_text} | "
                f"{row['health']} | {row['next_action']} |"
            )
    md_path.write_text("\n".join(markdown) + "\n", encoding="utf-8")

    state = StateStore(campaign_root)
    state.event("audit", runs=len(rows), destination=str(destination))
    state.artifact("audit_json", json_path)
    state.artifact("audit_csv", csv_path)
    state.artifact("audit_summary_csv", summary_csv_path)
    if workbook_written:
        state.artifact("audit_workbook", workbook_path)
    state.artifact("audit_markdown", md_path)
    payload["outputs"] = {
        "json": str(json_path),
        "csv": str(csv_path),
        "summary_csv": str(summary_csv_path),
        "markdown": str(md_path),
    }
    if workbook_written:
        payload["outputs"]["workbook"] = str(workbook_path)
    return payload
